from collections import Counter, defaultdict
from pathlib import Path


EMOTION_COLORS = {
    "neutral": "7F8C8D",
    "felicidad": "F1C40F",
    "tristeza": "3498DB",
    "enojo": "E74C3C",
    "miedo": "9B59B6",
    "disgusto": "27AE60",
    "sorpresa": "E67E22",
}


def _format_headers(sheet, headers):
    from openpyxl.styles import Font, PatternFill

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")


def _autosize_columns(sheet):
    for column in sheet.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 45)


def _format_elapsed_hms(elapsed_s: float) -> str:
    total_seconds = max(0, int(elapsed_s))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def _session_video_duration_s(session_data: dict, started_at, ended_at) -> float:
    video_fps = session_data.get("video_fps", 0.0)
    recorded_frames = session_data.get("recorded_frames", 0)
    if video_fps and video_fps > 0 and recorded_frames >= 0:
        return recorded_frames / video_fps
    return (ended_at - started_at).total_seconds()


def _emotion_message(emotion: str) -> str:
    messages = {
        "neutral": "Se observo una expresion estable durante gran parte de la sesion.",
        "felicidad": "Se observaron senales frecuentes de bienestar o estado positivo.",
        "tristeza": "Se detectaron momentos asociados a un estado de animo bajo.",
        "enojo": "Se observaron periodos con expresiones de tension o molestia.",
        "miedo": "Se detectaron expresiones relacionadas con alerta o incomodidad.",
        "disgusto": "Se observaron reacciones de rechazo en algunos momentos.",
        "sorpresa": "Se detectaron cambios de expresion asociados a sorpresa.",
    }
    return messages.get(emotion, "No hubo datos suficientes para una interpretacion general.")


def _presence_label(pct: float) -> str:
    if pct >= 50:
        return "alta"
    if pct >= 25:
        return "media"
    if pct > 0:
        return "baja"
    return "no observada"


def _build_timeline_sheet(workbook, events, title: str):
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet(title)

    emotions_seen = []
    for event in events:
        if event["top_1"] not in emotions_seen:
            emotions_seen.append(event["top_1"])
        if event["top_2"] not in emotions_seen:
            emotions_seen.append(event["top_2"])

    if not emotions_seen:
        emotions_seen = list(EMOTION_COLORS.keys())

    _format_headers(
        ws,
        ["Tiempo (mm:ss)", "Emocion principal", "Emocion secundaria", *emotions_seen],
    )

    for idx, event in enumerate(events, start=2):
        row_values = [
            _format_elapsed_hms(event["elapsed_s"])[3:],
            event["top_1"],
            event["top_2"],
        ]
        for emotion in emotions_seen:
            if event["top_1"] == emotion or event["top_2"] == emotion:
                row_values.append(emotions_seen.index(emotion))
            else:
                row_values.append(None)
        ws.append(row_values)

        color_1 = EMOTION_COLORS.get(event["top_1"], "34495E")
        color_2 = EMOTION_COLORS.get(event["top_2"], "34495E")
        ws[f"B{idx}"].fill = PatternFill(start_color=color_1, end_color=color_1, fill_type="solid")
        ws[f"B{idx}"].font = Font(color="FFFFFF")
        ws[f"C{idx}"].fill = PatternFill(start_color=color_2, end_color=color_2, fill_type="solid")
        ws[f"C{idx}"].font = Font(color="FFFFFF")

    if not events:
        ws.append(["-", "Sin datos", "Sin datos", *[0 for _ in emotions_seen]])
        _autosize_columns(ws)
        for col_idx in range(4, 4 + len(emotions_seen)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 1.2
        return

    legend_start = len(events) + 4
    ws[f"A{legend_start}"] = "Leyenda de codigos"
    ws[f"A{legend_start}"].font = Font(bold=True)

    for idx, emotion in enumerate(emotions_seen, start=0):
        row = legend_start + idx + 1
        color = EMOTION_COLORS.get(emotion, "34495E")
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = emotion
        ws[f"B{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"B{row}"].font = Font(color="FFFFFF")

    chart = LineChart()
    chart.title = "Linea de tiempo emocional (principal y secundaria)"
    chart.x_axis.title = "Tiempo (minuto:segundo)"
    chart.y_axis.title = "Codigo de emocion"
    chart.style = 10
    chart.legend.position = "r"
    chart.height = 12
    chart.width = 24
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    max_row = len(events) + 1
    data = Reference(ws, min_col=4, min_row=1, max_col=3 + len(emotions_seen), max_row=max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    for series, emotion_name in zip(chart.series, emotions_seen):
        color = EMOTION_COLORS.get(emotion_name, "34495E")
        series.graphicalProperties.line.noFill = True
        series.marker.symbol = "circle"
        series.marker.size = 6
        series.marker.graphicalProperties.solidFill = color
        series.marker.graphicalProperties.line.solidFill = color
        series.smooth = False

    chart.y_axis.scaling.min = -0.5
    chart.y_axis.scaling.max = max(len(emotions_seen) - 0.5, 0.5)
    chart.y_axis.majorUnit = 1

    ws.add_chart(chart, "E2")

    note_row = legend_start + len(emotions_seen) + 1
    ws[f"A{note_row}"] = (
        "Cada punto representa una emocion detectada en ese instante. "
        "Puede haber dos puntos por tiempo (emocion principal y emocion secundaria)."
    )
    ws[f"A{note_row}"].alignment = Alignment(wrap_text=True)

    _autosize_columns(ws)
    for col_idx in range(4, 4 + len(emotions_seen)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 1.2


def _add_terms_note_sheet(workbook, title: str = "Notas"):
    from openpyxl.styles import Alignment, Font

    ws = workbook.create_sheet(title)
    ws["A1"] = "Definiciones del reporte"
    ws["A1"].font = Font(bold=True, size=14)

    rows = [
        (
            "Emocion principal",
            "Es la emocion que se muestra con mas fuerza en ese punto de tiempo.",
        ),
        (
            "Emocion secundaria",
            "Es la segunda emocion mas presente en ese mismo punto, despues de la emocion principal.",
        ),
        (
            "Lectura sugerida",
            "Interpretar la evolucion completa de la sesion y no solo un instante aislado.",
        ),
    ]

    for idx, (label, value) in enumerate(rows, start=3):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].font = Font(bold=True)
        ws[f"B{idx}"] = value
        ws[f"B{idx}"].alignment = Alignment(wrap_text=True)

    _autosize_columns(ws)


def write_excel_report(report_path: Path, session_data: dict, ended_at):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as err:
        raise RuntimeError("Falta la dependencia 'openpyxl'. Instala con: pip install openpyxl") from err

    events = session_data["events"]
    started_at = session_data["started_at"]
    video_name = session_data["video_path"].name

    duration_s = _session_video_duration_s(session_data, started_at, ended_at)
    emotions_count = Counter(event["top_1"] for event in events)
    secondary_emotions_count = Counter(event["top_2"] for event in events)
    top1_conf_by_emotion = defaultdict(list)
    top2_conf_by_emotion = defaultdict(list)
    total_predictions = len(events)

    for event in events:
        top1_conf_by_emotion[event["top_1"]].append(event["top_1_conf"])
        top2_conf_by_emotion[event["top_2"]].append(event["top_2_conf"])

    dominant_emotion = "N/A"
    dominant_pct = 0.0
    secondary_dominant_emotion = "N/A"
    secondary_dominant_pct = 0.0
    avg_top1_conf = 0.0
    avg_top2_conf = 0.0
    avg_conf_gap = 0.0

    if emotions_count:
        dominant_emotion, dominant_count = emotions_count.most_common(1)[0]
        dominant_pct = (dominant_count / total_predictions) * 100.0
        avg_top1_conf = sum(event["top_1_conf"] for event in events) / total_predictions
        avg_top2_conf = sum(event["top_2_conf"] for event in events) / total_predictions
        avg_conf_gap = avg_top1_conf - avg_top2_conf

    if secondary_emotions_count:
        secondary_dominant_emotion, secondary_dominant_count = secondary_emotions_count.most_common(1)[0]
        secondary_dominant_pct = (secondary_dominant_count / total_predictions) * 100.0

    workbook = Workbook()

    ws_summary = workbook.active
    ws_summary.title = "Resumen"
    ws_summary["A1"] = "Informe de deteccion de emociones"
    ws_summary["A1"].font = Font(bold=True, size=14)

    summary_rows = [
        ("Fecha inicio", started_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Fecha fin", ended_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Duracion (segundos)", round(duration_s, 2)),
        ("Total de clasificaciones", total_predictions),
        ("Emocion predominante", dominant_emotion),
        ("Predominancia (%)", round(dominant_pct, 2)),
        ("Emocion secundaria mas frecuente", secondary_dominant_emotion),
        ("Predominancia secundaria (%)", round(secondary_dominant_pct, 2)),
        ("Confianza media emocion principal", round(avg_top1_conf, 4)),
        ("Confianza media emocion secundaria", round(avg_top2_conf, 4)),
        ("Brecha media de confianza (top1-top2)", round(avg_conf_gap, 4)),
        ("Archivo de video", video_name),
    ]

    for idx, (label, value) in enumerate(summary_rows, start=3):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"A{idx}"].font = Font(bold=True)
        ws_summary[f"B{idx}"] = value

    note_row = len(summary_rows) + 4
    ws_summary[f"A{note_row}"] = "Nota"
    ws_summary[f"A{note_row}"].font = Font(bold=True)
    ws_summary[f"B{note_row}"] = (
        "Este reporte es un apoyo descriptivo y no reemplaza una evaluacion "
        "clinica o psicosocial profesional."
    )
    ws_summary[f"B{note_row}"].alignment = Alignment(wrap_text=True)

    ws_predictions = workbook.create_sheet("Clasificaciones")
    _format_headers(
        ws_predictions,
        [
            "#",
            "Tiempo desde inicio (s)",
            "Tiempo desde inicio (HH:MM:SS)",
            "Emocion principal",
            "Confianza principal",
            "Emocion secundaria",
            "Confianza secundaria",
        ],
    )

    if not events:
        ws_predictions.append(["-", "-", "-", "Sin clasificaciones", "-", "-", "-"])
    else:
        for idx, event in enumerate(events, start=1):
            ws_predictions.append(
                [
                    idx,
                    round(event["elapsed_s"], 2),
                    _format_elapsed_hms(event["elapsed_s"]),
                    event["top_1"],
                    round(event["top_1_conf"], 4),
                    event["top_2"],
                    round(event["top_2_conf"], 4),
                ]
            )

    ws_distribution = workbook.create_sheet("Distribucion")
    _format_headers(
        ws_distribution,
        [
            "Tipo",
            "Emocion",
            "Frecuencia",
            "Porcentaje (%)",
            "Confianza promedio",
            "Confianza minima",
            "Confianza maxima",
        ],
    )

    if not events:
        ws_distribution.append(["Sin datos", "-", 0, 0.0, 0.0, 0.0, 0.0])
    else:
        for emotion, count in emotions_count.most_common():
            pct = (count / total_predictions) * 100.0
            confidences = top1_conf_by_emotion[emotion]
            avg_conf = sum(confidences) / len(confidences)
            min_conf = min(confidences)
            max_conf = max(confidences)
            ws_distribution.append(
                [
                    "Principal",
                    emotion,
                    count,
                    round(pct, 2),
                    round(avg_conf, 4),
                    round(min_conf, 4),
                    round(max_conf, 4),
                ]
            )
        for emotion, count in secondary_emotions_count.most_common():
            pct = (count / total_predictions) * 100.0
            confidences = top2_conf_by_emotion[emotion]
            avg_conf = sum(confidences) / len(confidences)
            min_conf = min(confidences)
            max_conf = max(confidences)
            ws_distribution.append(
                [
                    "Secundaria",
                    emotion,
                    count,
                    round(pct, 2),
                    round(avg_conf, 4),
                    round(min_conf, 4),
                    round(max_conf, 4),
                ]
            )

    _autosize_columns(ws_summary)
    _autosize_columns(ws_predictions)
    _autosize_columns(ws_distribution)
    _build_timeline_sheet(workbook, events, "LineaTiempo")
    _add_terms_note_sheet(workbook, "Notas")

    workbook.save(report_path)
    return report_path


def write_non_technical_excel_report(report_path: Path, session_data: dict, ended_at):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as err:
        raise RuntimeError("Falta la dependencia 'openpyxl'. Instala con: pip install openpyxl") from err

    events = session_data["events"]
    started_at = session_data["started_at"]
    video_name = session_data["video_path"].name

    duration_s = _session_video_duration_s(session_data, started_at, ended_at)
    total_predictions = len(events)
    emotions_count = Counter(event["top_1"] for event in events)
    secondary_emotions_count = Counter(event["top_2"] for event in events)

    dominant_emotion = "N/A"
    secondary_dominant_emotion = "N/A"
    dominant_pct = 0.0
    secondary_dominant_pct = 0.0

    if emotions_count:
        dominant_emotion, dominant_count = emotions_count.most_common(1)[0]
        dominant_pct = (dominant_count / total_predictions) * 100.0

    if secondary_emotions_count:
        secondary_dominant_emotion, secondary_dominant_count = secondary_emotions_count.most_common(1)[0]
        secondary_dominant_pct = (secondary_dominant_count / total_predictions) * 100.0

    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "ResumenSimple"
    ws_summary["A1"] = "Reporte no tecnico de emociones"
    ws_summary["A1"].font = Font(bold=True, size=14)

    summary_rows = [
        ("Inicio de sesion", started_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Fin de sesion", ended_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Duracion aproximada", _format_elapsed_hms(duration_s)),
        ("Emocion principal mas observada", dominant_emotion),
        (
            "Emocion secundaria mas observada",
            secondary_dominant_emotion,
        ),
        ("Nivel de presencia emocion principal", _presence_label(dominant_pct)),
        ("Nivel de presencia emocion secundaria", _presence_label(secondary_dominant_pct)),
        ("Interpretacion general", _emotion_message(dominant_emotion)),
        ("Archivo de video", video_name),
    ]

    for idx, (label, value) in enumerate(summary_rows, start=3):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"A{idx}"].font = Font(bold=True)
        ws_summary[f"B{idx}"] = value

    note_row = len(summary_rows) + 4
    ws_summary[f"A{note_row}"] = "Nota de uso"
    ws_summary[f"A{note_row}"].font = Font(bold=True)
    ws_summary[f"B{note_row}"] = (
        "Este reporte es informativo. Debe interpretarse junto con contexto "
        "pedagogico, clinico o psicosocial por personal profesional."
    )
    ws_summary[f"B{note_row}"].alignment = Alignment(wrap_text=True)

    _autosize_columns(ws_summary)
    _build_timeline_sheet(workbook, events, "LineaTiempoSimple")
    _add_terms_note_sheet(workbook, "Notas")
    workbook.save(report_path)
    return report_path
